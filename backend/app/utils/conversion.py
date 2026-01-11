"""File conversion utilities."""

from __future__ import annotations

import json
import subprocess
import tempfile
from pathlib import Path

import magic
from contextgem import DocxConverter  # pyright: ignore[reportMissingTypeStubs]
from openpyxl import load_workbook


class ConversionError(Exception):
    """File conversion failed."""


# MIME types the LLM accepts directly
LLM_SUPPORTED_TYPES = {
    "application/pdf",
    "image/png",
    "image/jpeg",
    "image/webp",
    "image/gif",
}

DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def to_pdf(data: bytes) -> bytes:
    """Convert DOCX/XLSX bytes to PDF."""
    mime = magic.from_buffer(data, mime=True)
    ext = {DOCX_MIME: ".docx", XLSX_MIME: ".xlsx"}.get(mime)
    if not ext:
        raise ConversionError(f"Cannot convert {mime} to PDF")

    with tempfile.TemporaryDirectory() as td:
        td_path = Path(td)
        input_path = td_path / f"input{ext}"
        input_path.write_bytes(data)
        return _libreoffice_convert(input_path, td_path)


def to_summary(data: bytes, mime: str) -> str:
    """Convert template to text summary for LLM context."""
    if mime == DOCX_MIME:
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "input.docx"
            path.write_bytes(data)
            return DocxConverter().convert_to_text_format(
                path, output_format="markdown"
            )

    if mime == XLSX_MIME:
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "input.xlsx"
            path.write_bytes(data)
            wb = load_workbook(path, data_only=False, read_only=True)
            return json.dumps(
                {
                    name: {
                        "columns": list(rows[0]) if rows else [],
                        "rows": [list(r) for r in rows[1:]],
                    }
                    for name in wb.sheetnames
                    for rows in [
                        list(wb[name].iter_rows(max_row=101, values_only=True))
                    ]
                },
                indent=2,
                default=str,
            )

    raise ConversionError(f"Cannot summarize {mime}")


def _libreoffice_convert(input_path: Path, output_dir: Path) -> bytes:
    """Run LibreOffice headless conversion to PDF."""
    import shutil
    import platform
    import uuid

    # Find soffice
    soffice = shutil.which("soffice")
    if not soffice and platform.system() == "Darwin":
        macos_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
        soffice = macos_path if Path(macos_path).exists() else "soffice"
    soffice = soffice or "soffice"

    profile = output_dir / f".lo_profile_{uuid.uuid4().hex[:8]}"

    try:
        result = subprocess.run(
            [
                soffice,
                "--headless",
                f"-env:UserInstallation=file://{profile}",
                "--convert-to",
                "pdf",
                "--outdir",
                str(output_dir),
                str(input_path),
            ],
            capture_output=True,
            text=True,
            timeout=60,
            check=False,
        )
    except subprocess.TimeoutExpired as e:
        raise ConversionError("LibreOffice timed out") from e
    except FileNotFoundError as e:
        raise ConversionError("LibreOffice not installed") from e

    if result.returncode != 0:
        raise ConversionError(f"LibreOffice failed: {result.stderr}")

    pdfs = list(output_dir.glob("*.pdf"))
    if not pdfs:
        raise ConversionError("No PDF output created")

    return pdfs[0].read_bytes()
