from __future__ import annotations

import io
import subprocess
import tempfile
from pathlib import Path
import json
from contextgem import DocxConverter
from openpyxl import load_workbook

from app.schemas.file_schemas import FileRead, FileStore
from typing import Literal


class ConversionError(Exception):
    """Raised when a file cannot be converted to PDF."""


IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
DOC_EXTS = {".pdf", ".docx", ".xlsx"}

OPEN_XML_EXTS = {".docx", ".xlsx"}


def file_type_to_mime_type(
    file_type: Literal["pdf", "docx", "xlsx", "png", "jpg", "jpeg", "webp", "gif"],
) -> str:
    if file_type == "docx":
        return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif file_type == "xlsx":
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif file_type == "pdf":
        return "application/pdf"
    elif (
        file_type == "png"
        or file_type == "jpg"
        or file_type == "jpeg"
        or file_type == "webp"
        or file_type == "gif"
    ):
        return f"image/{file_type}"
    else:
        raise ValueError(f"Unsupported ArtefactType: {file_type}")


def mime_type_to_file_type(mime_type: str) -> str:
    if (
        mime_type
        == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    ):
        return "docx"
    elif (
        mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ):
        return "xlsx"
    elif mime_type == "application/pdf":
        return "pdf"
    elif mime_type.startswith("image/"):
        return mime_type.split("/")[1]
    else:
        raise ValueError(f"Unsupported mime_type: {mime_type}")


def _get_soffice_path() -> str:
    """Get the path to the soffice executable, handling macOS installation location."""
    import shutil
    import platform

    # Check if soffice is in PATH
    soffice = shutil.which("soffice")
    if soffice:
        return soffice

    # On macOS, check the standard LibreOffice installation path
    if platform.system() == "Darwin":
        macos_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
        if Path(macos_path).exists():
            return macos_path

    # Fall back to "soffice" and let subprocess raise FileNotFoundError
    return "soffice"


def _run_libreoffice_conversion(input_path: Path, output_dir: Path) -> bytes:
    """Common LibreOffice conversion logic for converting documents to PDF."""
    soffice_path = _get_soffice_path()

    # Use a unique user profile to avoid conflicts with running LibreOffice instances
    import uuid

    user_installation = output_dir / f".libreoffice_profile_{uuid.uuid4().hex[:8]}"

    try:
        result = subprocess.run(
            [
                soffice_path,
                "--headless",
                f"-env:UserInstallation=file://{user_installation}",
                "--convert-to",
                "pdf",
                "--outdir",
                str(output_dir),
                str(input_path),
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=False,
            text=True,
            timeout=60,  # Prevent hanging on problematic files
        )
    except subprocess.TimeoutExpired as exc:
        raise ConversionError("LibreOffice conversion timed out") from exc
    except FileNotFoundError as exc:
        raise ConversionError(
            "soffice not found. Install LibreOffice: brew install --cask libreoffice (macOS) "
            "or apt-get install libreoffice-writer (Linux)"
        ) from exc
    except Exception as exc:
        raise ConversionError(f"Failed to launch soffice: {exc}") from exc

    if result.returncode != 0:
        raise ConversionError(
            f"LibreOffice conversion failed.\nstdout: {result.stdout}\nstderr: {result.stderr}"
        )

    pdf_candidates = list(output_dir.glob("*.pdf"))
    if not pdf_candidates:
        raise ConversionError(
            f"PDF output not created. LibreOffice may be running in GUI mode.\n"
            f"stdout: {result.stdout}\nstderr: {result.stderr}\n"
            f"Try: Close LibreOffice GUI or run 'pkill -f soffice'"
        )
    return pdf_candidates[0].read_bytes()


def _convert_docx_bytes(data: bytes) -> bytes:
    """Convert DOCX bytes to PDF bytes using LibreOffice."""
    with tempfile.TemporaryDirectory() as td:
        td_path = Path(td)
        in_path = td_path / "input.docx"
        in_path.write_bytes(data)
        return _run_libreoffice_conversion(in_path, td_path)


def _convert_xlsx_bytes(data: bytes) -> bytes:
    """Convert XLSX bytes to PDF bytes using LibreOffice."""
    with tempfile.TemporaryDirectory() as td:
        td_path = Path(td)
        in_path = td_path / "input.xlsx"
        in_path.write_bytes(data)
        return _run_libreoffice_conversion(in_path, td_path)


def _docx_to_md(data) -> str:
    """Convert a DOCX file to markdown or raw text using ContextGem's DocxConverter."""
    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "input.docx"
        path.write_bytes(data)
        converter = DocxConverter()
        docx_text = converter.convert_to_text_format(
            path,
            output_format="markdown",  # or "raw"
        )
        return docx_text


def _xlsx_to_json_str(data) -> str:
    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "input.xlsx"
        path.write_bytes(data)

        # read_only=True is essential for speed; data_only=False preserves formulas
        wb = load_workbook(path, data_only=False, read_only=True)
        result = {}

        for name in wb.sheetnames:
            # Fetch up to 101 rows (1 for header + 100 for data) in one go
            data = list(wb[name].iter_rows(max_row=101, values_only=True))

            result[name] = {
                "columns": list(data[0]) if data else [],
                "rows": [list(row) for row in data[1:]],  # The next 100 rows
            }

        # default=str ensures dates and other Excel objects don't break the JSON
        return json.dumps(result, indent=2, default=str)


def convert_to_pdf(file: FileStore) -> FileStore:
    """Convert a document (DOCX or XLSX) to PDF bytes."""
    if (
        file.mime_type
        == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    ):
        data = _convert_docx_bytes(file.data)
    elif (
        file.mime_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ):
        data = _convert_xlsx_bytes(file.data)
    else:
        raise ConversionError(
            f"Unsupported file extension for PDF conversion: {file.mime_type}"
        )
    return FileStore(
        storage_key=Path(file.storage_key).with_suffix(".pdf").as_posix(),
        file_name=Path(file.file_name).with_suffix(".pdf").name,
        mime_type="application/pdf",
        data=data,
        role=file.role,
    )


def get_template_summary(file: FileRead) -> str:
    """Convert a DOCX template to markdown summary string."""
    if (
        file.mime_type
        == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    ):
        return _docx_to_md(file.data)
    elif (
        file.mime_type
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ):
        return _xlsx_to_json_str(file.data)
    else:
        raise ConversionError(
            f"Unsupported file extension for template conversion: {file.mime_type}"
        )


def to_file_obj(file: FileRead) -> io.BytesIO:
    """Return a BytesIO with name set for OpenAI uploads."""
    bio = io.BytesIO(file.data)
    bio.name = file.file_name
    return bio
