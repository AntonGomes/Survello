from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlmodel import Session

from app.core.logging import logger
from app.models.dilaps_model import (
    DilapsItem,
    DilapsRun,
    DilapsSection,
    DilapsUnit,
)

HEADER_FONT = Font(name="Arial", bold=True, size=10)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14)
HEADER_FILL = PatternFill(start_color="D9E1F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
COLUMN_WIDTHS = {
    "A": 10,
    "B": 14,
    "C": 40,
    "D": 40,
    "E": 6,
    "F": 8,
    "G": 8,
    "H": 12,
}
HEADERS = [
    "Item No",
    "Lease clause",
    "Want of repair",
    "Remedy",
    "U",
    "Q",
    "R",
    "£",
]
PRELIMINARIES_RATE = 0.2
CONTRACT_ADMIN_RATE = 0.1
PRINCIPAL_DESIGNER_RATE = 0.01
NEGOTIATION_RATE = 0.05


def _write_headers(ws) -> None:
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def _write_section_header(ws, row: int, name: str) -> int:
    cell = ws.cell(row=row, column=3, value=name)
    cell.font = Font(name="Arial", bold=True, size=10)
    return row + 2


def _write_item_row(ws, row: int, item: DilapsItem) -> int:
    ws.cell(row=row, column=1, value=item.item_number).font = BODY_FONT
    ws.cell(row=row, column=2, value=item.lease_clause).font = BODY_FONT

    repair_cell = ws.cell(row=row, column=3, value=item.want_of_repair)
    repair_cell.font = BODY_FONT
    repair_cell.alignment = Alignment(wrap_text=True)

    remedy_cell = ws.cell(row=row, column=4, value=item.remedy)
    remedy_cell.font = BODY_FONT
    remedy_cell.alignment = Alignment(wrap_text=True)

    ws.cell(row=row, column=5, value=item.unit.value).font = BODY_FONT

    if item.unit != DilapsUnit.SUM and item.quantity is not None:
        ws.cell(row=row, column=6, value=float(item.quantity)).font = BODY_FONT
    if item.unit != DilapsUnit.SUM and item.rate is not None:
        ws.cell(row=row, column=7, value=float(item.rate)).font = BODY_FONT

    cost_cell = ws.cell(row=row, column=8)
    if item.quantity and item.rate and item.unit != DilapsUnit.SUM:
        cost_cell.value = f"=F{row}*G{row}"
    elif item.cost is not None:
        cost_cell.value = float(item.cost)
    cost_cell.font = BODY_FONT
    cost_cell.number_format = "#,##0"

    for col in range(1, 9):
        ws.cell(row=row, column=col).border = THIN_BORDER

    return row + 2


def _write_section_total(ws, row: int, first_row: int) -> int:
    cell = ws.cell(row=row, column=3, value="Total for Collection")
    cell.font = HEADER_FONT
    total = ws.cell(row=row, column=8, value=f"=SUM(H{first_row}:H{row - 1})")
    total.font = HEADER_FONT
    total.number_format = "#,##0"
    return row


def _write_front_pages(wb: Workbook, dilaps_run: DilapsRun) -> None:
    ws = wb.active
    ws.title = "Front Pages"

    ws.cell(row=2, column=1, value="DRAFT").font = TITLE_FONT
    ws.cell(row=4, column=1, value="Schedule of Dilapidations").font = TITLE_FONT
    ws.cell(row=6, column=1, value="of").font = BODY_FONT
    ws.cell(row=8, column=1, value=dilaps_run.property_address).font = TITLE_FONT

    if dilaps_run.lease_summary:
        clauses_label = "General clauses and preliminaries"
        ws.cell(row=14, column=1, value=clauses_label).font = HEADER_FONT
        summary_cell = ws.cell(row=16, column=1)
        summary_cell.value = dilaps_run.lease_summary
        summary_cell.font = BODY_FONT
        summary_cell.alignment = Alignment(wrap_text=True)


def _write_schedule_sheet(
    wb: Workbook,
    sheet_name: str,
    sections: list[DilapsSection],
) -> int:
    ws = wb.create_sheet(title=sheet_name[:31])
    _write_headers(ws)

    row = 3
    first_data_row = row
    for section in sections:
        row = _write_section_header(ws, row, section.name)
        for item in section.items:
            row = _write_item_row(ws, row, item)

    total_row = _write_section_total(ws, row, first_data_row)
    return total_row


def _write_collection_refs(ws, sheet_refs) -> tuple[int, int]:
    row = 5
    for sheet_name, total_row in sheet_refs:
        ws.cell(row=row, column=3, value=sheet_name).font = BODY_FONT
        ref = f"='{sheet_name}'!H{total_row}"
        ws.cell(row=row, column=8, value=ref).font = BODY_FONT
        ws.cell(row=row, column=8).number_format = "#,##0"
        row += 2

    sub_row = row
    ws.cell(row=sub_row, column=3, value="Sub total").font = HEADER_FONT
    formula = f"=SUM(H5:H{sub_row - 1})"
    ws.cell(row=sub_row, column=8, value=formula)
    ws.cell(row=sub_row, column=8).font = HEADER_FONT
    ws.cell(row=sub_row, column=8).number_format = "#,##0"

    row = sub_row + 2
    ws.cell(row=row, column=3, value="Contractor preliminaries").font = BODY_FONT
    ws.cell(row=row, column=4, value="Allow 20%").font = BODY_FONT
    ws.cell(row=row, column=8, value=f"=H{sub_row}*{PRELIMINARIES_RATE}")
    ws.cell(row=row, column=8).number_format = "#,##0"

    row += 2
    sub2_row = row
    ws.cell(row=row, column=3, value="Sub Total of work items").font = HEADER_FONT
    ws.cell(row=row, column=8, value=f"=SUM(H{sub_row}:H{row - 1})")
    ws.cell(row=row, column=8).font = HEADER_FONT
    ws.cell(row=row, column=8).number_format = "#,##0"
    return row, sub2_row


def _write_collection_fees(ws, row: int, sub2_row: int) -> None:
    row += 2
    fees_label = "Professional fees for managing the works"
    ws.cell(row=row, column=3, value=fees_label).font = HEADER_FONT

    row += 1
    ws.cell(row=row, column=3, value="Contract administration").font = BODY_FONT
    ws.cell(row=row, column=4, value="Allow 10%").font = BODY_FONT
    ws.cell(row=row, column=8, value=f"=H{sub2_row}*{CONTRACT_ADMIN_RATE}")
    ws.cell(row=row, column=8).number_format = "#,##0"

    row += 1
    ws.cell(row=row, column=3, value="Principal Designer").font = BODY_FONT
    ws.cell(row=row, column=4, value="Allow 1%").font = BODY_FONT
    formula = f"=H{sub2_row}*{PRINCIPAL_DESIGNER_RATE}"
    ws.cell(row=row, column=8, value=formula)
    ws.cell(row=row, column=8).number_format = "#,##0"

    row += 2
    fees_row = row
    ws.cell(row=row, column=3, value="Sub Total").font = HEADER_FONT
    ws.cell(row=row, column=8, value=f"=SUM(H{sub2_row}:H{row - 1})")
    ws.cell(row=row, column=8).font = HEADER_FONT
    ws.cell(row=row, column=8).number_format = "#,##0"

    row += 2
    prep_label = "Professional Fees (exc. VAT) for preparation"
    ws.cell(row=row, column=3, value=prep_label).font = BODY_FONT
    ws.cell(row=row, column=8, value=1750).font = BODY_FONT
    ws.cell(row=row, column=8).number_format = "#,##0"

    row += 2
    neg_label = "Professional Fees (exc. VAT) for Negotiation"
    ws.cell(row=row, column=3, value=neg_label).font = BODY_FONT
    ws.cell(row=row, column=4, value="Allow 5%").font = BODY_FONT
    ws.cell(row=row, column=8, value=f"=H{fees_row}*{NEGOTIATION_RATE}")
    ws.cell(row=row, column=8).number_format = "#,##0"

    row += 2
    ws.cell(row=row, column=3, value="GRAND TOTAL EXCLUDING VAT").font = HEADER_FONT
    ws.cell(row=row, column=8, value=f"=SUM(H{fees_row}:H{row - 1})")
    grand_font = Font(name="Arial", bold=True, size=12)
    ws.cell(row=row, column=8).font = grand_font
    ws.cell(row=row, column=8).number_format = "#,##0"


def _write_collection(
    wb: Workbook,
    sheet_refs: list[tuple[str, int]],
) -> None:
    ws = wb.create_sheet(title="Collection")

    ws.cell(row=1, column=1, value="Item No").font = HEADER_FONT
    ws.cell(row=1, column=8, value="£").font = HEADER_FONT

    ws.cell(row=3, column=3, value="COLLECTION").font = HEADER_FONT
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["H"].width = 14

    row, sub2_row = _write_collection_refs(ws, sheet_refs)
    _write_collection_fees(ws, row, sub2_row)


def export_dilaps(dilaps_run: DilapsRun, db: Session) -> bytes:
    logger.info(f"Exporting dilaps run {dilaps_run.id}")
    db.refresh(dilaps_run)
    sections = dilaps_run.sections

    wb = Workbook()
    _write_front_pages(wb, dilaps_run)

    sheet_refs: list[tuple[str, int]] = []
    section_threshold = 20

    if len(sections) <= section_threshold:
        total_row = _write_schedule_sheet(wb, "Schedule", sections)
        sheet_refs.append(("Schedule", total_row))
    else:
        for section in sections:
            name = (section.sheet_name or section.name)[:31]
            total_row = _write_schedule_sheet(wb, name, [section])
            sheet_refs.append((name, total_row))

    _write_collection(wb, sheet_refs)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
