from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Tuple

import openpyxl
import json
from openpyxl import load_workbook


def _norm_value(v: Any) -> Any:
    """Normalise values so trivial formatting differences do not break equality too often."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s
    return v


def _is_non_empty(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, str) and v.strip() == "":
        return False
    return True


def _load_workbook(path: Path) -> openpyxl.Workbook:
    if not path.exists():
        raise FileNotFoundError(path)
    # data_only=True compares displayed values (post-formula) if the file has cached results
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _sheet_names(wb: openpyxl.Workbook) -> set[str]:
    return set(wb.sheetnames)


def _occupied_column_header_names(wb: openpyxl.Workbook) -> set[str]:
    """
    Collect 'column names' from row 1 for columns that are occupied (have any non-empty cell)
    in any worksheet. If a column is occupied but has an empty header, we use a stable placeholder
    based on column index so structure can still match.
    """
    names: set[str] = set()

    for ws in wb.worksheets:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            continue

        for col in range(1, max_col + 1):
            occupied = False
            # Scan down the column until we find a non-empty cell
            for row in range(1, max_row + 1):
                v = ws.cell(row=row, column=col).value
                if _is_non_empty(v):
                    occupied = True
                    break

            if not occupied:
                continue

            header_v = ws.cell(row=1, column=col).value
            header = _norm_value(header_v)
            if _is_non_empty(header):
                names.add(str(header))
            else:
                names.add(f"__EMPTY_HEADER_COL_{col}")

    return names


def _non_empty_cell_map(wb: openpyxl.Workbook, only_sheets: set[str]) -> Dict[Tuple[str, int, int], Any]:
    """
    Map non-empty cells to their normalised value, restricted to selected sheet names.
    Key: (sheet_name, row, col)
    """
    out: Dict[Tuple[str, int, int], Any] = {}

    for name in wb.sheetnames:
        if name not in only_sheets:
            continue
        ws = wb[name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            continue

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                v = ws.cell(row=row, column=col).value
                if _is_non_empty(v):
                    out[(name, row, col)] = _norm_value(v)

    return out


def compare_spreadsheets(path_a: Path, path_b: Path) -> dict:
    """
    Measures similarity of 2 spreadsheets by averaging over these metrics:
      1) Identical column names in occupied columns
      2) Identical non-empty cells
      3) Identical worksheets (by name)

    Input: 2 Path values (one for each spreadsheet)
    Output: dict with raw counts + per-metric similarities + average similarity
    """
    wb_a = _load_workbook(path_a)
    wb_b = _load_workbook(path_b)

    # 3) Worksheets (by name)
    sheets_a = _sheet_names(wb_a)
    sheets_b = _sheet_names(wb_b)
    sheets_inter = sheets_a & sheets_b
    sheets_union = sheets_a | sheets_b

    worksheets_identical_count = len(sheets_inter)
    worksheets_union_count = len(sheets_union)
    worksheets_similarity = (worksheets_identical_count / worksheets_union_count) if worksheets_union_count else 1.0

    # 1) Identical column names in occupied columns
    cols_a = _occupied_column_header_names(wb_a)
    cols_b = _occupied_column_header_names(wb_b)
    cols_inter = cols_a & cols_b
    cols_union = cols_a | cols_b

    identical_column_names_count = len(cols_inter)
    occupied_column_names_union_count = len(cols_union)
    column_names_similarity = (identical_column_names_count / occupied_column_names_union_count) if occupied_column_names_union_count else 1.0

    # 2) Identical non-empty cells (restricted to common sheet names)
    cells_a = _non_empty_cell_map(wb_a, only_sheets=sheets_inter)
    cells_b = _non_empty_cell_map(wb_b, only_sheets=sheets_inter)

    keys_union = set(cells_a.keys()) | set(cells_b.keys())
    identical_non_empty_cells_count = 0
    for k in keys_union:
        if k in cells_a and k in cells_b and cells_a[k] == cells_b[k]:
            identical_non_empty_cells_count += 1

    non_empty_cells_union_count = len(keys_union)
    non_empty_cells_similarity = (identical_non_empty_cells_count / non_empty_cells_union_count) if non_empty_cells_union_count else 1.0

    # Average similarity across the 3 metrics
    average_similarity = (column_names_similarity + non_empty_cells_similarity + worksheets_similarity) / 3.0

    return {
        "paths": {"a": str(path_a), "b": str(path_b)},
        "worksheets": {
            "identical_by_name": worksheets_identical_count,
            "union_count": worksheets_union_count,
            "similarity": worksheets_similarity,
        },
        "column_names_in_occupied_columns": {
            "identical_names": identical_column_names_count,
            "union_count": occupied_column_names_union_count,
            "similarity": column_names_similarity,
        },
        "non_empty_cells": {
            "identical_cells": identical_non_empty_cells_count,
            "union_count": non_empty_cells_union_count,
            "similarity": non_empty_cells_similarity,
            "note": "Compared only within worksheets present in both files, and only for non-empty cells.",
        },
        "average_similarity": average_similarity,
    }

def excel_to_json(path: Path) -> str:
    # read_only=True is essential for speed; data_only=False preserves formulas
    wb = load_workbook(path, data_only=False, read_only=True)
    result = {}

    for name in wb.sheetnames:
        # Fetch up to 101 rows (1 for header + 100 for data) in one go
        data = list(wb[name].iter_rows(max_row=101, values_only=True))
        
        result[name] = {
            "columns": list(data[0]) if data else [],
            "rows": [list(row) for row in data[1:]] # The next 100 rows
        }
    
    # default=str ensures dates and other Excel objects don't break the JSON
    return json.dumps(result, indent=2, default=str)
